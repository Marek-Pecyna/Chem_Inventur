from openpyxl import load_workbook
import csv
from datetime import datetime
import xlsxwriter


class Model:
    def __init__(self, filename=None, delimiter=None, encoding=None, primary_key=None, dialect=csv.excel):
        self.filename = filename
        self.delimiter = delimiter
        self.encoding = encoding
        self.primary_key = primary_key
        self.dialect = dialect
        self.database = {}
        self.column_names = []
        if filename:
            self.get_data_from_csv_file(filename=filename,
                                        delimiter=delimiter,
                                        encoding=encoding,
                                        primary_key=primary_key,
                                        dialect=dialect)
        print(f'{__name__}:      Model wurde initialisiert.')
        return

    def import_excel_file(self, filename: str, sheet_name: str = None):
        """
        From an Excel file (from old stock-taking ("Inventur"))
        if no sheet_name is given, active sheet will be processed
        """
        print(f"\n{__name__}:      Analyse Excel-Datei begonnen.")
        self.filename = filename
        # Load your workbook and sheets as you want, for example
        wb = load_workbook(filename)
        header = []
        db = {}  # dictionary with all data, keys = consecutive number

        if sheet_name:
            actual_sheet = wb[sheet_name]
        else:
            actual_sheet = wb[wb.active.title]

        print(f"   \u2022 Arbeitsblatt '{actual_sheet}' wird bearbeitet")
        data = list(actual_sheet.iter_rows(values_only=True))  # read in
        print(f"   \u2022 {len(data)} Einträge mit Header")

        for index, row in enumerate(data):
            if index == 0:  # analysis of header
                print("   \u2022 Header enthält folgende Spalten: ",
                      end="")
                for entry in row:
                    if entry:
                        header.append(entry)
                    else:
                        header.append('')
                    print(f"{entry}", end=",")
                print()
            else:
                """
                row_db = {'Name': None,
                          'Alternativer Name 1': None,
                          'Alternativer Name 2': None,
                          'Englischer Name': None,
                          'Alter des Gebindes': None,
                          'Reinheitsgrad': None,
                          'Zustandsform': None,
                          'Verwendungszweck': None,
                          'Standort': None,
                          'Summenformel': None,
                          'CAS-Nummer': None,
                          'Gebindegröße': None,
                          'Restmenge': None,
                          'Hersteller': None,
                          'Bestellnummer': None,
                          'Gefahrenklasse': None,
                          'H-Sätze': None,
                          'P-Sätze': None,
                          'Link SDB': None,
                          'Link GESTIS': None,
                          'Letzte Aktualisierung': None,
                          }
                """
                row_db = {}
                for position, entry in enumerate(row):
                    # print(header[position])
                    if entry:
                        row_db.update({header[position]: entry})
                    else:
                        row_db.update({header[position]: ''})
                db.update({index: row_db})
        self.column_names = header
        self.database = db
        for key in db:
            print(key, db[key])
            break

        print(f"{__name__}:      Analyse beendet.\n")
        return

    def save_as_csv_file(self, filename: str):
        print()
        with open(filename, 'w', newline='', encoding='utf-16') as csvfile:
            writer = csv.writer(csvfile, dialect=self.dialect, delimiter=self.delimiter)
            writer.writerow(self.column_names)
            # writing all lines
            for key in self.database:
                text = []
                for feature in self.database[key]:
                    text.append(self.database[key][feature])
                writer.writerow(text)
        print(f"{__name__}:      {datetime.now().strftime('%d.%m.%Y %H:%M:%S')} CSV-Datei '{filename}' wurde "
              f"gespeichert.\n")
        return

    def export_as_excel_file(self, filename: str):
        print()
        # Create workbook
        workbook = xlsxwriter.Workbook(filename)
        header_cell_format = workbook.add_format({'bg_color': r'#d4ddcf', 'font_size': 12, 'bottom': True})
        integer_format = workbook.add_format()
        integer_format.set_num_format('0; [Red] (-0); [Magenta] 0')

        worksheet = workbook.add_worksheet("Inventur Biochemie 2023")
        worksheet.freeze_panes(1, 2)
        worksheet.set_zoom(100)
        worksheet.set_column(0, 0, 5)
        worksheet.set_column("B:Y", 25)
        worksheet.autofilter(0, 0, len(self.database), len(self.column_names))

        row = 0
        col = 0
        for item in self.column_names:
            worksheet.write(row, col, item, header_cell_format)
            col += 1

        for key in self.database:
            row += 1
            col = 0
            for feature in self.database[key]:
                worksheet.write(row, col, self.database[key][feature])
                col += 1

        workbook.close()
        print(f"{__name__}:      {datetime.now().strftime('%d.%m.%Y %H:%M:%S')} Excel-Datei '{filename}' wurde gespeichert.\n")
        return

    def get_data_from_csv_file(self,
                               filename: str,
                               delimiter=';',
                               encoding='utf-16',
                               primary_key=None,
                               dialect=None):
        self.filename = filename
        self.delimiter = delimiter
        self.encoding = encoding
        print(f"\n{__name__}:      Analyse CSV-Datei begonnen.")

        with open(filename, newline='', encoding=encoding) as csv_file:
            # Create CSV reader
            csv_reader = csv.DictReader(csv_file, delimiter=delimiter, dialect=dialect)

            # Extract fieldnames without primary key
            self.column_names = csv_reader.fieldnames
            print(f'{__name__}:      Spaltennamen: {self.column_names}')
            if primary_key:
                self.primary_key = primary_key
            else:
                self.primary_key = self.column_names[0]
            print(f"{__name__}:      Primary key: '{self.primary_key}'")

            # Extract data from reader
            for row in csv_reader:
                tmp = {}
                for index, key in enumerate(row):
                    # use value as primary key
                    if key == self.primary_key:
                        # get integer values for primary key if possible:
                        try:
                            prim_key = int(row[key])
                        except ValueError:
                            prim_key = row(key)
                        tmp.update({key: prim_key})
                        continue
                    tmp.update({key: row[key]})
                    # Create entry in dictionary
                # only if 'Name' is empty:
                if 'Name' in tmp and tmp['Name'] == "":
                    if 'Handelsname' in tmp:
                        tmp['Name'] = tmp['Handelsname']
                    else:
                        tmp['Name'] = prim_key
                self.database.update({prim_key: tmp})
        print(f"{__name__}:      Analyse beendet.\n")
        return
